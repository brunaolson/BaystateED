#!/usr/bin/env python3
"""
parse_pem_data.py
-----------------
Parses the PEM Staffing Summit Excel workbook and exports clean, long-format
CSV files to /Users/brunaolson/BaystateED/data/.

Source file : /Users/brunaolson/Downloads/PEM Staffing Summit Data_3.22.26 (2).xlsx
Data period : FY 2024/25  (Oct 2024 – Mar 2026  /  Apr 2024 – Sept 2025)

Run:  python3 parse_pem_data.py

Outputs (all in the same directory as this script):
  hourly_census_winter.csv          – FW sheet, long format
  hourly_census_summer.csv          – SS sheet, long format
  hourly_census_raw_winter.csv      – Oct-Mar sheet, long format (full precision)
  hourly_census_raw_summer.csv      – Apr-Sept sheet, long format (full precision)
  coverage_hours.csv                – shift schedules + provider variability
  provider_variability.csv          – MD/APP/Resident hour percentiles
  staffing_analysis_winter.csv      – FW Analysis hourly attending capacity
  staffing_analysis_summer.csv      – SS Analysis hourly attending capacity
  capacity_28day.csv                – 28-day rotation provider assignments
  resident_shifts.csv               – Sheet3 daily resident shift counts
  graph_data_summer.csv             – Graph Data SS section
  graph_data_winter.csv             – Graph Data FW section
"""

import csv
import os
import datetime
import openpyxl

# ── paths ────────────────────────────────────────────────────────────────────
EXCEL_PATH = "/Users/brunaolson/Downloads/PEM Staffing Summit Data_3.22.26 (2).xlsx"
OUT_DIR = os.path.dirname(os.path.abspath(__file__))   # same dir as this script

# ── helpers ──────────────────────────────────────────────────────────────────

def metric_label(raw):
    """Normalise metric identifiers to clean strings."""
    if raw == 0.25:
        return "p25"
    if raw == 0.75:
        return "p75"
    if raw == 0.95:
        return "p95"
    if raw == 0.1:
        return "p10"
    if raw == 0.9:
        return "p90"
    if raw == 0.5:
        return "p50"
    if isinstance(raw, str):
        return raw.strip().lower().replace(" ", "_")
    return str(raw)


CATEGORY_SLUG = {
    "Total Pedi ED Census":                   "census",
    "WR Census":                              "wr",
    "Crisis/Bedsearch Boarding Census":       "boarding",
    "ED Hourly New Arrival":                  "arrivals",
    "ED Hourly New Bed Assignment":           "bed_assignment",
    "ED Hourly Checkout (Discharged Pts)":    "discharge",
    "ED Hourly Checkout (Admitted Pts)":      "admit",
    "Count of LWBS based on Arrival Times":   "lwbs",
}


def slugify_category(raw):
    return CATEGORY_SLUG.get(raw, raw)


def out(filename):
    return os.path.join(OUT_DIR, filename)


# ── 1. FW / SS clean summary sheets ──────────────────────────────────────────

def parse_summary_sheet(ws):
    """
    Parse FW or SS sheet.  Structure:
        Row 1  : header  (Category | Metric | 0:00 | 1:00 … 23:00)
        Rows 2+: data rows with category, metric, then 24 hour values
    Returns list of dicts: season, category, metric, hour, value
    """
    rows = list(ws.iter_rows(values_only=True))
    # row 0 is header; columns 2..25 are hours 0..23
    records = []
    for row in rows[1:]:   # skip header
        category_raw = row[0]
        metric_raw   = row[1]
        if category_raw is None or metric_raw is None:
            continue
        category = slugify_category(str(category_raw).strip())
        metric   = metric_label(metric_raw)
        for hour in range(24):
            val = row[2 + hour]
            if val is None:
                continue
            records.append({
                "category": category,
                "metric":   metric,
                "hour":     hour,
                "value":    val,
            })
    return records


# ── 2. Oct-Mar / Apr-Sept raw sheets ─────────────────────────────────────────

RAW_METRIC_ORDER = ["Mean", "Median", 0.25, 0.75, 0.95, "STDEV"]

def parse_raw_sheet(ws):
    """
    Parse Oct-Mar or Apr-Sept.  Structure:
        Row 1  : title + hour columns (datetime.time objects)
        Row 2  : category header (no metrics)
        Rows 3-8: metric rows (6 per category)
        Row 9  : next category header …  etc.
    Returns list of dicts: category, metric, hour, value
    """
    rows = list(ws.iter_rows(values_only=True))

    # Determine current category by scanning col A
    records = []
    current_category = None

    for row in rows[1:]:       # skip title row
        col_a = row[0]
        col_b = row[1]         # first hour value (hour 0)

        # Row is a category header when col_b is None and col_a is a string
        if col_a is not None and col_b is None:
            current_category = slugify_category(str(col_a).strip())
            continue

        # Row is a metric data row when col_a is a known metric indicator
        if col_a in RAW_METRIC_ORDER or (isinstance(col_a, float) and col_a in (0.25, 0.75, 0.95)):
            if current_category is None:
                continue
            metric = metric_label(col_a)
            for hour in range(24):
                val = row[1 + hour]   # col B = hour 0, col C = hour 1 …
                if val is None:
                    continue
                records.append({
                    "category": current_category,
                    "metric":   metric,
                    "hour":     hour,
                    "value":    round(float(val), 6) if isinstance(val, float) else val,
                })

    return records


# ── 3. Coverage Hours ─────────────────────────────────────────────────────────

def parse_coverage_hours(ws):
    """
    Returns list of dicts for the shift schedule tables.
    Columns: season, shift_name, day_of_week, hours
    """
    rows = list(ws.iter_rows(values_only=True))

    records = []
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    # Oct-March block: rows 11-18 (0-indexed 10-17), header at row 11 (index 11)
    # April-Sept block: rows 21-27 (0-indexed 20-26)

    SEASON_BLOCKS = [
        ("Oct-Mar",   10, 17),   # row indices (0-based) inclusive
        ("Apr-Sept",  20, 26),
    ]

    for season, r_start, r_end in SEASON_BLOCKS:
        header_row = rows[r_start + 1]   # "Shift", "Monday" … "Sunday"
        for ri in range(r_start + 2, r_end + 1):
            row = rows[ri]
            shift_name = row[0]
            if shift_name is None or shift_name == "TOTAL":
                continue
            for di, day in enumerate(days):
                val = row[1 + di]
                if val is None:
                    continue
                records.append({
                    "season":     season,
                    "shift_name": shift_name,
                    "day_of_week": day,
                    "hours":       val,
                })

    return records


# ── 4. Provider Variability ───────────────────────────────────────────────────

def parse_provider_variability(ws):
    """
    From Coverage Hours rows 1-4:
      Category | P10 | P25 | P50 | P75 | P90
    Returns list of dicts: provider_type, percentile, hours
    """
    rows = list(ws.iter_rows(values_only=True))
    percentiles = ["p10", "p25", "p50", "p75", "p90"]
    records = []
    for row in rows[1:4]:   # rows 2-4 (0-indexed 1-3)
        ptype = row[0]
        if ptype is None:
            continue
        for i, pct in enumerate(percentiles):
            val = row[1 + i]
            if val is None:
                continue
            records.append({
                "provider_type": ptype,
                "percentile":    pct,
                "hours":         val,
            })
    return records


# ── 5. FW Analysis / SS Analysis ─────────────────────────────────────────────

def parse_analysis_sheet(ws, season):
    """
    Parse FW Analysis or SS Analysis.
    The hourly data starts after the shift schedule block.
    For FW: column offset 13 (0-indexed), hour col = col 13, shifts in cols 14-18, SUM=19, CAP=20
    For SS: column offset 18 (0-indexed), hour col = 18, shifts in cols 19-22, SUM=23, CAP=24

    Returns list of dicts: season, hour, shift_name, attending_shifts, capacity_pph
    """
    rows = list(ws.iter_rows(values_only=True))

    # Find the row that has "Attending shifts per day:"
    header_row_idx = None
    for i, row in enumerate(rows):
        if any(isinstance(v, str) and "Attending shifts per day" in v for v in row):
            header_row_idx = i
            break

    if header_row_idx is None:
        return []

    # The column that contains the hour label is the same column as the "Attending shifts per day" label
    label_col = None
    for ci, v in enumerate(rows[header_row_idx]):
        if isinstance(v, str) and "Attending shifts per day" in v:
            label_col = ci
            break

    shift_names_row = rows[header_row_idx + 1]   # shift time strings
    # Shift name columns start at label_col+1
    # Detect actual shift name strings in rows header+1 (times like "7:30a-4p")
    shift_col_names = {}
    for ci in range(label_col + 1, len(shift_names_row)):
        v = shift_names_row[ci]
        if v is None:
            continue
        if isinstance(v, str) and (v.strip() in ("SUM", "CAPACITY", "PPH") or v.strip() == ""):
            break
        shift_col_names[ci] = str(v).strip()

    # Also look at the row above (header_row_idx) for shift labels like "Day", "Evening"
    label_names_row = rows[header_row_idx]
    shift_col_labels = {}
    for ci in range(label_col + 1, len(label_names_row)):
        v = label_names_row[ci]
        if v is None:
            continue
        if isinstance(v, str) and v.strip() in ("SUM", "CAPACITY", ""):
            break
        shift_col_labels[ci] = str(v).strip()

    # Merge: prefer the label from the header_row (Day/Evening/etc.) supplemented by time from next row
    shift_cols = {}
    all_cols = set(list(shift_col_labels.keys()) + list(shift_col_names.keys()))
    for ci in sorted(all_cols):
        label = shift_col_labels.get(ci, "")
        time  = shift_col_names.get(ci, "")
        if label and time:
            shift_cols[ci] = f"{label} ({time})"
        elif label:
            shift_cols[ci] = label
        elif time:
            shift_cols[ci] = time

    # Find SUM and CAPACITY column indices
    sum_col = None
    cap_col = None
    for ci, v in enumerate(rows[header_row_idx]):
        if isinstance(v, str):
            if v.strip() == "SUM":
                sum_col = ci
            if v.strip() == "CAPACITY":
                cap_col = ci

    # Get PPH value (look for it near the shift schedule block)
    pph = 2.5   # default from both sheets

    # Hourly data rows: starting 2 rows after the header row (skip the time-strings row)
    records = []
    data_start = header_row_idx + 2
    for row in rows[data_start:]:
        hour_val = row[label_col]
        if hour_val is None:
            continue
        try:
            hour = int(hour_val)
        except (TypeError, ValueError):
            continue
        if not (0 <= hour <= 23):
            continue

        # Per-shift attending fractions
        for ci, shift_name in shift_cols.items():
            v = row[ci]
            if v is None:
                v = 0.0
            records.append({
                "season":      season,
                "hour":        hour,
                "shift_name":  shift_name,
                "attending_shifts": v,
                "capacity_pph": pph,
            })

        # Also record the sum and total capacity
        sum_val = row[sum_col] if sum_col is not None else None
        cap_val = row[cap_col] if cap_col is not None else None
        records.append({
            "season":      season,
            "hour":        hour,
            "shift_name":  "TOTAL",
            "attending_shifts": sum_val,
            "capacity_pph": cap_val,
        })

    return records


# ── 6. Total Capacity Analysis (28-day rotation) ─────────────────────────────

PROVIDER_CODES = {"EMS", "EMI", "AP", "F", "Pi1", "Pi2", "Ps1", "Ps2"}


def parse_capacity_28day(ws):
    """
    Extracts:
      a) The 28-day rotation grid: day, slot_position (1-N), provider_type
         Layout: row 3 (0-based row 2) has the first day-header using cols 2-7;
                 subsequent day-headers use cols 1-7.
      b) The hourly capacity table from cols 10-22.
    Returns two lists of dicts.
    """
    rows = list(ws.iter_rows(values_only=True))

    # ---- a) Rotation grid ----
    # Strategy:
    #   - Scan every row in the grid area (rows 2-76, 0-indexed).
    #   - A "day header" row is one whose cells in cols 1-7 contain only ints or None,
    #     with at least one int.
    #   - Between two day-header rows, all rows that contain at least one PROVIDER_CODE
    #     in the active day columns are "provider rows".
    #   - First group (days 1-6) has day ints in cols 2-7 (col 1 is None);
    #     all subsequent groups have day ints in cols 1-7.

    rotation_records = []

    def is_day_header(row):
        vals = [row[c] for c in range(1, 8)]
        return (all(isinstance(v, (int, type(None))) for v in vals)
                and any(isinstance(v, int) for v in vals))

    def has_provider(row, col_start, num_days):
        for ci in range(col_start, col_start + num_days):
            v = row[ci]
            if isinstance(v, str) and v.strip() in PROVIDER_CODES:
                return True
        return False

    groups = []    # list of (days_list, col_start, provider_rows_list)
    current_days = None
    current_col_start = None
    current_num_days = None
    current_prows = []

    for row in rows[2:77]:    # rows 3-77 in 1-based Excel
        if is_day_header(row):
            # flush previous group
            if current_days:
                groups.append((current_days, current_col_start, current_prows))
            # determine layout for this group
            if row[1] is None:
                # first block: days in cols 2-7
                days_list = [row[c] for c in range(2, 8)]
                col_start = 2
            else:
                days_list = [row[c] for c in range(1, 8)]
                col_start = 1
            days_list = [d for d in days_list if isinstance(d, int)]
            current_days = days_list
            current_col_start = col_start
            current_num_days = len(days_list)
            current_prows = []
        elif current_days is not None:
            if has_provider(row, current_col_start, current_num_days):
                current_prows.append(row)

    # flush final group
    if current_days:
        groups.append((current_days, current_col_start, current_prows))

    # flatten groups into records
    for days_list, col_start, prows in groups:
        for slot_idx, prow in enumerate(prows):
            for day_idx, day in enumerate(days_list):
                val = prow[col_start + day_idx]
                if val is not None and str(val).strip() in PROVIDER_CODES:
                    rotation_records.append({
                        "day":           day,
                        "slot_position": slot_idx + 1,
                        "provider_type": str(val).strip(),
                    })

    # ---- b) Hourly capacity table ----
    # Located in rows 3-31 (0-indexed 2-30), cols 10-20
    # Header (row 2): Day | Junior_Hours | CAPACITY | Senior_Hours | CAPACITY |
    #                 Fellow_Hours | CAPACITY | APP_Hours | CAPACITY | TOTAL | total_capacity_per_9h
    # Col 21-22 contain a PPH legend (Junior/Senior/Fellow/APP rates) – not per-day data, excluded.
    capacity_records = []
    for row in rows[3:31]:
        day_val = row[10]
        if day_val is None:
            continue
        try:
            day = int(day_val)
        except (TypeError, ValueError):
            continue
        capacity_records.append({
            "day":                       day,
            "junior_hours":              row[11],
            "junior_capacity":           row[12],
            "senior_hours":              row[13],
            "senior_capacity":           row[14],
            "fellow_hours":              row[15],
            "fellow_capacity":           row[16],
            "app_hours":                 row[17],
            "app_capacity":              row[18],
            "total_provider_hours":      row[19],
            "total_patient_capacity_9h": row[20],
        })

    return rotation_records, capacity_records


# ── 7. Sheet3 – resident shifts ───────────────────────────────────────────────

def parse_resident_shifts(ws):
    """Returns list of dicts: day, junior_shifts, senior_shifts, junior_hours,
    senior_hours, total_shifts, total_hours.  Also includes provider summary rows."""
    rows = list(ws.iter_rows(values_only=True))
    shift_records = []
    summary_records = []

    for row in rows[1:29]:   # rows 2-29
        day = row[0]
        if not isinstance(day, int):
            continue
        shift_records.append({
            "day":             day,
            "junior_shifts":   row[1],
            "senior_shifts":   row[2],
            "junior_hours":    row[3],
            "senior_hours":    row[4],
            "total_shifts":    row[5],
            "total_hours":     row[6],
        })

    # Provider summary table: col 9=category, 10=avg_shifts, 11=avg_hours, 12=pct_coverage
    # rows 17-22 in Excel (0-indexed 16-21); row 17 is the column header row – skip it.
    for row in rows[17:22]:   # rows 18-22 are data
        category = row[9]
        if category is None or not isinstance(category, str):
            continue
        # Skip the header row if it slipped through
        if category.strip().lower() in ("provider category", "category"):
            continue
        summary_records.append({
            "provider_category":    category,
            "avg_shifts_per_day":   row[10],
            "avg_hours_per_day":    row[11],
            "pct_of_total_coverage": row[12],
        })

    return shift_records, summary_records


# ── 8. Graph Data ─────────────────────────────────────────────────────────────

def parse_graph_data(ws):
    """
    The sheet has two sections: SS (rows 1-15) and FW (rows 17-31).
    Structure per section: category | metric | hour_0 … hour_23
    Delta rows (category=None, metric="Delta 25%" etc.) are included as special category.
    Returns two lists of dicts (ss_records, fw_records).
    """
    rows = list(ws.iter_rows(values_only=True))

    def parse_block(block_rows, season):
        records = []
        current_category = None
        for row in block_rows:
            cat_raw = row[0]
            met_raw = row[1]
            if met_raw is None:
                continue
            if cat_raw is not None:
                current_category = slugify_category(str(cat_raw).strip()) if cat_raw not in ("SS", "FW") else None
            # Delta rows
            if isinstance(met_raw, str) and "delta" in met_raw.lower():
                metric_str = met_raw.strip().lower().replace(" ", "_")
                cat_str = current_category or "unknown"
            else:
                metric_str = metric_label(met_raw)
                cat_str = current_category or "unknown"
            for hour in range(24):
                val = row[2 + hour]
                if val is None:
                    continue
                records.append({
                    "season":   season,
                    "category": cat_str,
                    "metric":   metric_str,
                    "hour":     hour,
                    "value":    val,
                })
        return records

    # SS block: rows 0-14 (skip the header row 0 which has "SS" label)
    ss_records = parse_block(rows[1:15], "summer")
    # FW block: rows 16-30 (row 16 has "FW" label, skip it)
    fw_records = parse_block(rows[17:31], "winter")

    return ss_records, fw_records


# ── write helpers ─────────────────────────────────────────────────────────────

def write_csv(filename, fieldnames, records):
    path = out(filename)
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)
    print(f"  wrote {len(records):>6} rows  ->  {path}")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # 1. Hourly census – clean summary (FW / SS)
    print("\n[1/12] Hourly census – winter (FW sheet)")
    recs = parse_summary_sheet(wb["FW"])
    write_csv("hourly_census_winter.csv",
              ["category", "metric", "hour", "value"], recs)

    print("[2/12] Hourly census – summer (SS sheet)")
    recs = parse_summary_sheet(wb["SS"])
    write_csv("hourly_census_summer.csv",
              ["category", "metric", "hour", "value"], recs)

    # 2. Hourly census – raw precision (Oct-Mar / Apr-Sept)
    print("[3/12] Hourly census – raw winter (Oct-Mar sheet)")
    recs = parse_raw_sheet(wb["Oct-Mar"])
    write_csv("hourly_census_raw_winter.csv",
              ["category", "metric", "hour", "value"], recs)

    print("[4/12] Hourly census – raw summer (Apr-Sept sheet)")
    recs = parse_raw_sheet(wb["Apr-Sept"])
    write_csv("hourly_census_raw_summer.csv",
              ["category", "metric", "hour", "value"], recs)

    # 3. Coverage hours (shift schedules)
    print("[5/12] Coverage hours")
    recs = parse_coverage_hours(wb["Coverage Hours"])
    write_csv("coverage_hours.csv",
              ["season", "shift_name", "day_of_week", "hours"], recs)

    # 4. Provider variability
    print("[6/12] Provider variability")
    recs = parse_provider_variability(wb["Coverage Hours"])
    write_csv("provider_variability.csv",
              ["provider_type", "percentile", "hours"], recs)

    # 5. Staffing analysis
    print("[7/12] Staffing analysis – winter (FW Analysis sheet)")
    recs = parse_analysis_sheet(wb["FW Analysis"], "winter")
    write_csv("staffing_analysis_winter.csv",
              ["season", "hour", "shift_name", "attending_shifts", "capacity_pph"], recs)

    print("[8/12] Staffing analysis – summer (SS Analysis sheet)")
    recs = parse_analysis_sheet(wb["SS Analysis"], "summer")
    write_csv("staffing_analysis_summer.csv",
              ["season", "hour", "shift_name", "attending_shifts", "capacity_pph"], recs)

    # 6. 28-day capacity rotation + hourly capacity table
    print("[9/12] 28-day capacity rotation")
    rotation_recs, capacity_recs = parse_capacity_28day(wb["Total Capacity Analysis"])
    write_csv("capacity_28day.csv",
              ["day", "slot_position", "provider_type"], rotation_recs)
    write_csv("capacity_28day_hourly.csv",
              ["day", "junior_hours", "junior_capacity", "senior_hours", "senior_capacity",
               "fellow_hours", "fellow_capacity", "app_hours", "app_capacity",
               "total_provider_hours", "total_patient_capacity_9h"], capacity_recs)

    # 7. Resident shifts
    print("[10/12] Resident shifts (Sheet3)")
    shift_recs, summary_recs = parse_resident_shifts(wb["Sheet3"])
    write_csv("resident_shifts.csv",
              ["day", "junior_shifts", "senior_shifts", "junior_hours",
               "senior_hours", "total_shifts", "total_hours"], shift_recs)
    write_csv("provider_summary.csv",
              ["provider_category", "avg_shifts_per_day", "avg_hours_per_day",
               "pct_of_total_coverage"], summary_recs)

    # 8. Graph data
    print("[11/12] Graph data – summer")
    ss_graph, fw_graph = parse_graph_data(wb["Graph Data"])
    write_csv("graph_data_summer.csv",
              ["season", "category", "metric", "hour", "value"], ss_graph)

    print("[12/12] Graph data – winter")
    write_csv("graph_data_winter.csv",
              ["season", "category", "metric", "hour", "value"], fw_graph)

    print("\nAll files written successfully.")


if __name__ == "__main__":
    main()
